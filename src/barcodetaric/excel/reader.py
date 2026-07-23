"""Import κωδικολογίου πελάτη από Excel/CSV με αυτόματη ανίχνευση στηλών.

Ο χρήστης δίνει ένα αρχείο με barcode + περιγραφή (και προαιρετικά TARIC). Ο reader
εντοπίζει τις στήλες από τις κεφαλίδες ή, ελλείψει κεφαλίδων, από το περιεχόμενο.
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path

from ..engine.http_util import contains_greek

BARCODE_HEADERS = ["barcode", "μπαρκοντ", "μπαρκωδ", "ean", "upc", "κωδικος", "κωδικός", "code"]
DESC_HEADERS = ["desc", "περιγραφ", "ονομασ", "προιον", "προϊον", "product", "name", "είδος", "ειδος"]
TARIC_HEADERS = ["taric", "ταρικ", "hs", "δασμολ"]


@dataclass
class ImportedRow:
    barcode: str = ""
    description: str = ""
    taric_code: str = ""


def _clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _looks_like_barcode(value: str) -> bool:
    digits = re.sub(r"\D", "", value)
    return len(digits) in (8, 12, 13, 14) and len(digits) >= len(value.strip()) - 2


def _find_col(header: list[str], needles: list[str]) -> int | None:
    for i, h in enumerate(header):
        if any(n in h for n in needles):
            return i
    return None


def _read_rows(path: Path) -> list[list[str]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        sheet = wb.worksheets[0]
        return [[("" if c is None else str(c)) for c in row]
                for row in sheet.iter_rows(values_only=True)]
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            try:
                delimiter = csv.Sniffer().sniff(fh.read(2048), delimiters=",;\t").delimiter
                fh.seek(0)
            except csv.Error:
                fh.seek(0)
            return [list(r) for r in csv.reader(fh, delimiter=delimiter)]
    if suffix == ".txt":
        return [[line] for line in path.read_text(encoding="utf-8").splitlines()]
    raise ValueError(f"Μη υποστηριζόμενος τύπος αρχείου: {suffix}")


def read_codebook(path: str | Path) -> list[ImportedRow]:
    path = Path(path)
    rows = [r for r in _read_rows(path) if any(_clean(c) for c in r)]
    if not rows:
        return []

    header = [_clean(c).lower() for c in rows[0]]
    barcode_idx = _find_col(header, BARCODE_HEADERS)
    desc_idx = _find_col(header, DESC_HEADERS)
    taric_idx = _find_col(header, TARIC_HEADERS)
    has_header = barcode_idx is not None or desc_idx is not None or taric_idx is not None
    data = rows[1:] if has_header else rows

    # Χωρίς κεφαλίδες: μάντεψε στήλες από το περιεχόμενο.
    if barcode_idx is None and desc_idx is None:
        barcode_idx, desc_idx = _guess_columns(data)

    out: list[ImportedRow] = []
    for r in data:
        barcode = _clean(r[barcode_idx]) if barcode_idx is not None and barcode_idx < len(r) else ""
        desc = _clean(r[desc_idx]) if desc_idx is not None and desc_idx < len(r) else ""
        taric = _clean(r[taric_idx]) if taric_idx is not None and taric_idx < len(r) else ""
        if not desc and not barcode:
            continue
        # Αν λείπει η περιγραφή, πάρε την πρώτη μη-barcode στήλη κειμένου.
        if not desc:
            desc = _first_text(r, exclude={barcode_idx})
        out.append(ImportedRow(barcode=re.sub(r"\s", "", barcode), description=desc,
                               taric_code=re.sub(r"\D", "", taric)))
    return out


def _guess_columns(data: list[list[str]]) -> tuple[int | None, int | None]:
    ncols = max((len(r) for r in data[:30]), default=0)
    barcode_idx = None
    for i in range(ncols):
        hits = sum(1 for r in data[:30] if i < len(r) and _looks_like_barcode(_clean(r[i])))
        if hits >= max(2, len(data[:30]) // 3):
            barcode_idx = i
            break
    desc_idx = None
    best_len = 0
    for i in range(ncols):
        if i == barcode_idx:
            continue
        avg = sum(len(_clean(r[i])) for r in data[:30] if i < len(r)) / max(1, len(data[:30]))
        if avg > best_len:
            best_len = avg
            desc_idx = i
    return barcode_idx, desc_idx


def _first_text(row: list[str], exclude: set) -> str:
    for i, cell in enumerate(row):
        if i in exclude:
            continue
        val = _clean(cell)
        if val and not _looks_like_barcode(val):
            return val
    return ""
