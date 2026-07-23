"""Import της επίσημης ΕΕ ονοματολογίας (Combined Nomenclature) σε τοπική βάση.

Πηγή: data.europa.eu «Combined Nomenclature <έτος>» — δημοσιεύεται ετησίως,
πολύγλωσσα (με Ελληνικά), σε XML/CSV/XLSX. Δεν υπάρχει δωρεάν επίσημο REST API,
γι' αυτό κατεβάζουμε + κάνουμε parse τοπικά (offline matching μετά το import).

Ο importer είναι format-flexible ώστε να αντέχει τις παραλλαγές των distributions:
δέχεται τοπικό αρχείο ή URL, και ανιχνεύει αυτόματα .xml / .csv / .tsv / .xlsx / .zip.
Αν δεν υπάρχει καμία πηγή, φορτώνει ένα ενσωματωμένο seed (λίγες εγγραφές) ώστε το
app να είναι λειτουργικό από την πρώτη στιγμή.
"""

from __future__ import annotations

import csv
import io
import re
import tempfile
import urllib.request
import zipfile
from pathlib import Path
from typing import Callable, Iterable, Optional

from ..engine.http_util import BROWSER_UA, contains_greek, debug
from ..models import TaricRow
from .. import repo

# Σελίδα-πηγή (ενημερώνεται από updates.py). Το πραγματικό distribution URL
# αλλάζει ανά έτος· ο χρήστης μπορεί να το ορίσει ή να κάνει import τοπικό αρχείο.
DEFAULT_SOURCE_PAGE = "https://data.europa.eu/data/datasets/combined-nomenclature-2026"


# ------------------------------------------------------------- seed dataset ----
# Ελάχιστο ενσωματωμένο δείγμα (HS4 headings) για offline λειτουργία/δοκιμές.
# code, description_el, description_en
_SEED: list[tuple[str, str, str]] = [
    ("2201", "Νερά, συμπεριλαμβανομένων των φυσικών ή τεχνητών μεταλλικών νερών", "Waters, including natural or artificial mineral waters"),
    ("2202", "Νερά με προσθήκη ζάχαρης ή γλυκαντικών, αναψυκτικά", "Waters with added sugar/sweetener, soft drinks"),
    ("2203", "Μπίρα από βύνη", "Beer made from malt"),
    ("2204", "Κρασιά από νωπά σταφύλια", "Wine of fresh grapes"),
    ("0901", "Καφές", "Coffee"),
    ("1806", "Σοκολάτα και άλλα παρασκευάσματα με κακάο", "Chocolate and other cocoa preparations"),
    ("1905", "Προϊόντα αρτοποιίας, μπισκότα", "Bread, pastry, cakes, biscuits"),
    ("3304", "Καλλυντικά προϊόντα ή προϊόντα μακιγιάζ", "Beauty or make-up preparations"),
    ("3305", "Παρασκευάσματα για τα μαλλιά", "Preparations for use on the hair"),
    ("3401", "Σαπούνια", "Soap"),
    ("4818", "Χαρτί υγείας, χαρτομάντιλα, χειροπετσέτες", "Toilet paper, tissues, towels of paper"),
    ("2404", "Προϊόντα για εισπνοή χωρίς καύση· προϊόντα νικοτίνης", "Products for inhalation without combustion; nicotine products"),
    ("6109", "T-shirts και φανελάκια, πλεκτά", "T-shirts, singlets, knitted"),
    ("6203", "Κοστούμια, σακάκια, παντελόνια για άνδρες", "Men's suits, jackets, trousers"),
    ("8517", "Τηλέφωνα και συσκευές επικοινωνίας", "Telephones and communication apparatus"),
    ("8471", "Αυτόματες μηχανές επεξεργασίας δεδομένων (υπολογιστές)", "Automatic data-processing machines (computers)"),
]


def _seed_rows() -> list[TaricRow]:
    return [TaricRow(code=c, level=4, parent_code=c[:2], description_el=el, description_en=en,
                     hs4=c, indent=0, source_version="seed")
            for c, el, en in _SEED]


# ------------------------------------------------------------------ parsing ----

def _clean(text: object) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def _looks_like_code(value: str) -> bool:
    digits = re.sub(r"\D", "", value)
    return 2 <= len(digits) <= 10 and len(digits) % 2 == 0


def _to_row(code: str, desc_el: str, desc_en: str, version: str) -> TaricRow:
    digits = re.sub(r"\D", "", code)
    level = len(digits)
    return TaricRow(code=digits, level=level, parent_code=digits[:-2] if level > 2 else "",
                    description_el=desc_el, description_en=desc_en, hs4=digits[:4],
                    indent=0, source_version=version)


def _parse_tabular(rows: Iterable[list[str]], version: str) -> list[TaricRow]:
    """Ανιχνεύει στήλη κωδικού + στήλη(ες) περιγραφής από πίνακα (CSV/XLSX)."""
    rows = [r for r in rows if any(_clean(c) for c in r)]
    if not rows:
        return []

    header = [_clean(c).lower() for c in rows[0]]
    code_idx = _find_col(header, ["cn", "code", "κωδικ", "κωδικός", "goods code"])
    el_idx = _find_col(header, ["el", "greek", "ελλην", "περιγραφ"])
    en_idx = _find_col(header, ["en", "english", "description", "self-explanatory"])
    data_rows = rows[1:] if (code_idx is not None or en_idx is not None) else rows

    # Αν δεν βρέθηκαν headers, μάντεψε: πρώτη στήλη που μοιάζει με κωδικό.
    if code_idx is None:
        for i in range(len(rows[0])):
            if any(_looks_like_code(_clean(r[i])) for r in data_rows[:20] if i < len(r)):
                code_idx = i
                break
    if code_idx is None:
        return []

    out: list[TaricRow] = []
    for r in data_rows:
        if code_idx >= len(r):
            continue
        code = _clean(r[code_idx])
        if not _looks_like_code(code):
            continue
        el = _clean(r[el_idx]) if el_idx is not None and el_idx < len(r) else ""
        en = _clean(r[en_idx]) if en_idx is not None and en_idx < len(r) else ""
        # Αν έχουμε μία μόνο στήλη περιγραφής, ανίχνευσε γλώσσα από το περιεχόμενο.
        if not el and not en:
            desc = _first_text_cell(r, exclude=code_idx)
            el, en = (desc, "") if contains_greek(desc) else ("", desc)
        # Διόρθωση: αν το περιεχόμενο δεν ταιριάζει με τη στήλη-γλώσσα, μετακίνησέ το.
        if en and contains_greek(en) and not el:
            el, en = en, ""
        elif el and not contains_greek(el) and not en:
            el, en = "", el
        out.append(_to_row(code, el, en, version))
    return out


def _first_text_cell(row: list[str], exclude: int) -> str:
    for i, cell in enumerate(row):
        if i == exclude:
            continue
        val = _clean(cell)
        if len(val) > 2 and not _looks_like_code(val):
            return val
    return ""


def _find_col(header: list[str], needles: list[str]) -> Optional[int]:
    for i, h in enumerate(header):
        if any(n in h for n in needles):
            return i
    return None


def _parse_csv(path: Path, version: str) -> list[TaricRow]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(4096)
        fh.seek(0)
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t").delimiter
        except csv.Error:
            pass
        reader = csv.reader(fh, delimiter=delimiter)
        return _parse_tabular(list(reader), version)


def _parse_xlsx(path: Path, version: str) -> list[TaricRow]:
    from openpyxl import load_workbook
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    rows: list[list[str]] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            rows.append([("" if c is None else str(c)) for c in row])
    return _parse_tabular(rows, version)


def _parse_xml(path: Path, version: str) -> list[TaricRow]:
    """Γενικό XML parse: βρίσκει ζεύγη code/description ανεξαρτήτως ακριβούς schema."""
    import xml.etree.ElementTree as ET

    try:
        tree = ET.parse(path)
    except ET.ParseError as exc:
        debug(f"XML parse error: {exc}")
        return []
    root = tree.getroot()

    def _localname(tag: str) -> str:
        return tag.rsplit("}", 1)[-1].lower()

    out: list[TaricRow] = []
    # Ψάξε elements που περιέχουν παιδί με "code" και "description".
    for elem in root.iter():
        code = ""
        el_desc = ""
        en_desc = ""
        for child in list(elem):
            ln = _localname(child.tag)
            text = _clean(child.text)
            if not text:
                continue
            if "code" in ln and _looks_like_code(text):
                code = text
            elif "descr" in ln or "text" in ln or "label" in ln:
                lang = (child.get("lang") or child.get("{http://www.w3.org/XML/1998/namespace}lang") or "").lower()
                if lang.startswith("el") or contains_greek(text):
                    el_desc = text
                elif lang.startswith("en") or not en_desc:
                    en_desc = text
        if code and (el_desc or en_desc):
            out.append(_to_row(code, el_desc, en_desc, version))
    return out


def _parse_zip(path: Path, version: str) -> list[TaricRow]:
    out: list[TaricRow] = []
    with zipfile.ZipFile(path) as zf:
        for name in zf.namelist():
            suffix = Path(name).suffix.lower()
            if suffix not in {".xml", ".csv", ".tsv", ".xlsx"}:
                continue
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                tmp.write(zf.read(name))
                tmp_path = Path(tmp.name)
            try:
                out.extend(parse_file(tmp_path, version))
            finally:
                tmp_path.unlink(missing_ok=True)
    return out


def _indent_depth(indent: object) -> int:
    """'- - - ' -> 3. None/'' -> 0."""
    s = str(indent or "")
    return s.count("-")


def _struct_depth(code10: str, indent_depth: int) -> int:
    """Δομικό βάθος ιεραρχίας: κεφάλαιο(0)/κλάση(1)/υποκλάσεις(2+indent)."""
    if len(code10) >= 10 and code10[2:] == "00000000":
        return 0
    if len(code10) >= 10 and code10[4:] == "000000":
        return 1
    return 2 + indent_depth


def _parse_nomenclature_sheet(path: Path) -> dict[str, dict]:
    """Parse ενός CIRCABC «Nomenclature XX.xlsx». Χτίζει ιεραρχικό path.

    Επιστρέφει {goods_full: {code10, hs4, desc, path, indent_depth, valid_to}}.
    goods_full = π.χ. '0101210000 80' (κωδικός + product-line suffix).
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    out: dict[str, dict] = {}
    levels: dict[int, str] = {}
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if not row or not row[0]:
            continue
        goods_full = str(row[0]).strip()               # '0101210000 80'
        code10 = re.sub(r"\D", "", goods_full)[:10]
        if len(code10) < 2:
            continue
        indent = row[5] if len(row) > 5 else ""
        desc = _clean(row[6]) if len(row) > 6 else ""
        valid_to = _clean(row[2]) if len(row) > 2 else ""
        if not desc:
            continue
        depth = _struct_depth(code10, _indent_depth(indent))
        levels[depth] = desc
        for d in [k for k in levels if k > depth]:
            del levels[d]
        path = " > ".join(levels[k] for k in sorted(levels))
        out[goods_full] = {"code10": code10, "hs4": code10[:4], "desc": desc, "path": path,
                           "indent_depth": _indent_depth(indent), "valid_to": valid_to}
    wb.close()  # απελευθέρωση του αρχείου (Windows lock σε read_only mode)
    return out


def parse_nomenclature(el_path: Path, en_path: Path | None, version: str) -> list[TaricRow]:
    """Συγχωνεύει τα «Nomenclature EL/EN.xlsx» του CIRCABC σε TaricRow (με path context)."""
    el = _parse_nomenclature_sheet(el_path)
    en = _parse_nomenclature_sheet(en_path) if en_path else {}
    rows: list[TaricRow] = []
    for goods_full, e in el.items():
        en_e = en.get(goods_full, {})
        code10 = e["code10"]
        rows.append(TaricRow(
            code=code10, level=len(code10.rstrip("0")) or 2,
            parent_code=code10[:-2] if len(code10) > 2 else "",
            description_el=e["desc"], description_en=en_e.get("desc", ""),
            description_path_el=e["path"], description_path_en=en_e.get("path", ""),
            hs4=e["hs4"], indent=e["indent_depth"], valid_to=e["valid_to"],
            source_version=version,
        ))
    # Αν το EN έχει κωδικούς που λείπουν από το EL, πρόσθεσέ τους.
    for goods_full, en_e in en.items():
        if goods_full not in el:
            code10 = en_e["code10"]
            rows.append(TaricRow(code=code10, hs4=en_e["hs4"], description_en=en_e["desc"],
                                 description_path_en=en_e["path"], indent=en_e["indent_depth"],
                                 valid_to=en_e["valid_to"], source_version=version))
    return rows


def parse_file(path: Path, version: str = "") -> list[TaricRow]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        return _parse_csv(path, version)
    if suffix in {".xlsx", ".xlsm"}:
        return _parse_xlsx(path, version)
    if suffix == ".xml":
        return _parse_xml(path, version)
    if suffix == ".zip":
        return _parse_zip(path, version)
    raise ValueError(f"Μη υποστηριζόμενος τύπος αρχείου: {suffix}")


# --------------------------------------------------------------- public API ----

def import_from_file(path: str | Path, *, version: str = "", source_url: str = "",
                     progress: Optional[Callable[[str], None]] = None) -> int:
    path = Path(path)
    version = version or f"file:{path.name}"
    if progress:
        progress(f"Ανάλυση αρχείου {path.name}…")
    rows = parse_file(path, version)
    if not rows:
        raise ValueError("Δεν βρέθηκαν έγκυρες εγγραφές TARIC στο αρχείο.")
    if progress:
        progress(f"Εισαγωγή {len(rows)} εγγραφών…")
    return repo.bulk_insert_taric(rows, version=version, source_url=source_url or str(path))


def import_from_url(url: str, *, version: str = "",
                    progress: Optional[Callable[[str], None]] = None) -> int:
    if progress:
        progress(f"Λήψη από {url}…")
    suffix = Path(url.split("?")[0]).suffix.lower() or ".bin"
    req = urllib.request.Request(url, headers={"User-Agent": BROWSER_UA})
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = resp.read()
    # Ανίχνευση ZIP από magic bytes αν το URL δεν έχει κατάληξη.
    if data[:2] == b"PK":
        suffix = ".zip"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(data)
        tmp_path = Path(tmp.name)
    try:
        return import_from_file(tmp_path, version=version or f"url:{url}",
                                source_url=url, progress=progress)
    finally:
        tmp_path.unlink(missing_ok=True)


def import_seed(progress: Optional[Callable[[str], None]] = None) -> int:
    """Φορτώνει το ενσωματωμένο δείγμα — για offline λειτουργία/δοκιμές."""
    if progress:
        progress("Φόρτωση ενσωματωμένου δείγματος ΕΕ ονοματολογίας…")
    rows = _seed_rows()
    return repo.bulk_insert_taric(rows, version="seed", source_url="builtin-seed")
