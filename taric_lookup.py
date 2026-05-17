#!/usr/bin/env python3
"""Barcode/description to TARIC matcher with free/no-key defaults."""

from __future__ import annotations

import argparse
import csv
import html
import json
import os
import re
import sys
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, List, Optional

# Import full EU TARIC catalog with 160+ entries
try:
    from full_taric_catalog import FULL_TARIC_CATALOG, TaricEntry as FullTaricEntry
    HAS_FULL_CATALOG = True
except ImportError:
    HAS_FULL_CATALOG = False

# Import product database module
try:
    from product_database import (
        init_database, add_product, search_by_barcode, search_by_description,
        get_all_products, export_database_to_json, import_database_from_json,
        ProductRecord, DB_PATH
    )
    HAS_DATABASE = True
except ImportError:
    HAS_DATABASE = False

OPENFOODFACTS_URL = "https://world.openfoodfacts.org/api/v0/product/{barcode}.json"
POLLINATIONS_URL = "https://text.pollinations.ai/{prompt}"
OPENROUTER_URL = "https://openrouter.ai/api/v1/chat/completions"
DUCKDUCKGO_AI_URL = os.getenv("DUCKDUCKGO_AI_URL", "https://duckduckgo.com/duckchat/v1/chat")
DEEPINFRA_URL = os.getenv("DEEPINFRA_URL", "https://api.deepinfra.com/v1/openai/chat/completions")
DEEPINFRA_MODEL = os.getenv("DEEPINFRA_MODEL", "meta-llama/Meta-Llama-3.1-8B-Instruct-Turbo")
NERVE_URL = os.getenv("NERVE_URL", "")
UPCITEMDB_URL = "https://api.upcitemdb.com/prod/trial/lookup?upc={barcode}"
BARCODE_MONSTER_URL = "https://barcode.monster/api/{barcode}"
BARCODELOOKUP_URL = "https://www.barcodelookup.com/{barcode}"
GOUNPC_SEARCH_URL = "https://go-upc.com/search?q={barcode}"
EAN_SEARCH_URL = "https://www.ean-search.org/?q={barcode}"
WAVEGROCERY_PRODUCTS_SEARCH_URL = os.getenv(
    "WAVEGROCERY_PRODUCTS_SEARCH_URL",
    "https://thanopoulos.api.wavegrocery.com/api/v3.1/products/search?term={barcode}",
)
DOTENV_PATH = Path(".env")


@dataclass(frozen=True)
class TaricEntry:
    taric_code: str
    hs4: str
    description: str
    keywords: tuple[str, ...]


# Use full EU catalog if available, otherwise fallback to minimal catalog
if HAS_FULL_CATALOG:
    DEFAULT_TARIC_CATALOG = FULL_TARIC_CATALOG
else:
    # Fallback minimal catalog if full_taric_catalog.py not available
    DEFAULT_TARIC_CATALOG: tuple[TaricEntry, ...] = (
        # --- Chapter 22: Beverages ---
        TaricEntry(
            "2201101100", "2201",
            "Natural mineral waters, non-carbonated, not flavoured",
            ("natural mineral water", "still mineral water", "mineral water",
             "still water", "spring water", "water", "mineral", "drink"),
        ),
        TaricEntry(
            "2201101900", "2201",
            "Natural mineral waters, carbonated, not flavoured",
            ("sparkling mineral water", "carbonated mineral water", "aerated water",
             "sparkling water", "fizzy water"),
        ),
        TaricEntry(
            "2202100000", "2202",
            "Waters with added sugar, sweeteners or flavouring",
            ("flavoured water", "sweetened water", "water drink"),
        ),
        TaricEntry(
            "2202990000", "2202",
            "Other non-alcoholic beverages",
            ("energy drink", "soft drink", "soda", "cola", "lemonade", "isotonic", "monster", "red bull"),
        ),
        TaricEntry(
            "2203000000", "2203",
            "Beer made from malt",
            ("beer", "ale", "lager", "pilsner", "stout"),
        ),
        TaricEntry(
            "2204210000", "2204",
            "Wine of fresh grapes, not sparkling, not fortified, <= 2L",
            ("wine", "red wine", "white wine", "rose wine"),
        ),
        # --- Chapter 8: Electrical Machinery - Primary Batteries ---
        TaricEntry(
            "8506100010", "8506",
            "Alkaline manganese dioxide batteries",
            ("alkaline battery", "aa battery", "aaa battery", "primary battery", 
             "primary cell", "disposable battery", "zinc carbon battery"),
        ),
        TaricEntry(
            "8506600010", "8506",
            "Other primary batteries and cells",
            ("battery", "primary cell", "dry cell", "disposable cell"),
        ),
        # --- Chapter 24: Tobacco / Vaping ---
        TaricEntry(
            "2404120000", "2404",
            "Products for electronic cigarettes and similar devices, without nicotine",
            ("vape", "e-liquid", "electronic cigarette", "e-cigarette", "vaping",
             "liquid", "flavor shot", "flavour shot", "inhalation", "puff", "bar juice"),
        ),
        TaricEntry(
            "2404110000", "2404",
            "Products for electronic cigarettes and similar devices, with nicotine",
            ("nicotine e-liquid", "nic salt", "nicotine vape", "vape nicotine"),
        ),
        TaricEntry(
            "2402200000", "2402",
            "Cigarettes containing tobacco",
            ("cigarette", "tobacco cigarette", "smoking", "marlboro"),
        ),
        # --- Chapter 61/62: Clothing ---
        TaricEntry(
            "6105100000", "6105",
            "Men's shirts of cotton, knitted or crocheted",
            ("shirt", "mens", "cotton", "garment", "clothing", "t-shirt", "polo", "top"),
        ),
        TaricEntry(
            "6109100000", "6109",
            "T-shirts of cotton",
            ("t-shirt", "tshirt", "top", "cotton top", "singlet"),
        ),
        TaricEntry(
            "6203421100", "6203",
            "Men's trousers of denim (jeans)",
            ("jeans", "denim", "pants", "trousers", "denim pants"),
        ),
        TaricEntry(
            "6203420000", "6203",
            "Men's trousers of cotton",
            ("trousers", "pants", "cotton pants"),
        ),
        TaricEntry(
            "6402990500", "6402",
            "Sports footwear (Sneakers)",
            ("sneakers", "nike", "adidas", "shoes", "sports shoes", "athletic shoes"),
        ),
        # --- Chapter 84/85: Electronics ---
        TaricEntry(
            "8471300000", "8471",
            "Portable automatic data processing machines (Laptops)",
            ("laptop", "notebook", "computer", "macbook", "portable computer"),
        ),
        TaricEntry(
            "8517130000", "8517",
            "Smartphones (New Nomenclature)",
            ("smartphone", "iphone", "samsung", "android phone"),
        ),
        TaricEntry(
            "8517140000", "8517",
            "Other wireless mobile phones",
            ("mobile phone", "cell phone", "feature phone", "phone"),
        ),
        TaricEntry(
            "8518300000", "8518",
            "Headphones and earphones",
            ("headphones", "airpods", "earbuds", "earphones", "headset"),
        ),
        TaricEntry(
            "8542310000", "8542",
            "Electronic integrated circuits",
            ("chip", "semiconductor", "integrated circuit", "microchip", "cpu"),
        ),
        TaricEntry(
            "8504400000", "8504",
            "Power supply units and transformers",
            ("power adapter", "charger", "transformer", "power supply"),
        ),
        # --- Chapter 33: Cosmetics ---
        TaricEntry(
            "3304990000", "3304",
            "Beauty or make-up preparations",
            ("makeup", "cosmetic", "foundation", "lipstick", "mascara", "skincare", "cream"),
        ),
        TaricEntry(
            "3305100000", "3305",
            "Shampoos",
            ("shampoo", "hair wash", "hair shampoo"),
        ),
        TaricEntry(
            "3307200000", "3307",
            "Personal deodorants and antiperspirants",
            ("deodorant", "antiperspirant", "body spray"),
        ),
        # --- Chapter 30: Pharmaceuticals ---
        TaricEntry(
            "3004900000", "3004",
            "Medicaments for therapeutic use",
            ("medicine", "drug", "pill", "tablet", "capsule", "pharmaceutical", "paracetamol"),
        ),
        TaricEntry(
            "2106909200", "2106",
            "Food supplements (Vitamins and minerals)",
            ("vitamin", "supplement", "multivitamin", "dietary supplement"),
        ),
        # --- Chapter 95: Toys & Games ---
        TaricEntry(
            "9503000000", "9503",
            "Toys and puzzles",
            ("toy", "game", "puzzle", "doll", "action figure", "board game", "lego"),
        ),
        TaricEntry(
            "9504500000", "9504",
            "Video game consoles and equipment",
            ("playstation", "xbox", "nintendo", "console", "video game", "switch"),
        ),
    )


_TRANSLATIONS = {
    "ανδρ": "mens",
    "γυναικ": "womens",
    "παιδικ": "kids",
    "βαμβακερ": "cotton",
    "πουκαμισ": "shirt",
    "μπλουζ": "t-shirt",
    "παντελον": "trousers",
    "τζιν": "jeans",
    "μακρυμανικ": "long sleeve",
    "κοντομανικ": "short sleeve",
    "παπουτσ": "sneakers",
    "αθλητικ": "sports shoes",
    "λαπτοπ": "laptop",
    "φορητος υπολογιστης": "laptop",
    "υπολογιστης": "computer",
    "κινητο": "smartphone",
    "τηλεφων": "phone",
    "smartphone": "smartphone",
    "ακουστικ": "headphones",
    "κονσολα": "console",
    "νερο": "water",
    "μεταλλικο νερο": "mineral water",
    "ανθρακουχο": "carbonated",
    "αναψυκτικ": "soft drink",
    "ενεργειακο ποτο": "energy drink",
    "χυμο": "juice",
    "μπυρα": "beer",
    "κρασι": "wine",
    "σαμπουαν": "shampoo",
    "καλλυντικ": "cosmetic",
    "αποσμητικ": "deodorant",
    "φαρμακ": "medicine",
    "συμπληρωμα διατροφης": "supplement",
    "βιταμιν": "vitamin",
    "πολυβιταμιν": "multivitamin",
    "πολυβιταμινη": "multivitamin",
    "ηλεκτρονικο τσιγαρο": "electronic cigarette",
    "τσιγαρο": "cigarette",
    "υγρο vape": "vape e-liquid",
    "υγρο ηλεκτρονικου τσιγαρου": "vape e-liquid",
    "νικοτιν": "nicotine",
    "nic salt": "nicotine salt",
    "παιχνιδ": "toy",
    "σφραγιδα": "stamp",
    "μελανι": "ink",
    "μελάνι": "ink",
    "σφραγίδα": "stamp",
    "trodat": "ink",
    "αρωμα": "perfume",
    "κολονια": "cologne",
    "άρωμα": "perfume",
    "κολόνια": "cologne",
    "χαρτι υγειας": "toilet paper",
    "χαρτί υγείας": "toilet paper",
    "χαρτι": "paper",
    "χαρτί": "paper",
    "υγειας": "toilet",
    "υγείας": "toilet",
    "ρολο": "roll",
    "ρολό": "roll",
    "φυλλων": "ply",
    "φύλλων": "ply",
    "4φυλλο": "4-ply",
    "4φύλλο": "4-ply",
}


def _debug(message: str) -> None:
    if os.getenv("BARCODE_TARIC_DEBUG") == "1":
        print(f"[debug] {message}", file=sys.stderr)


def _get_effective_ai_provider(provider: str) -> str:
    if provider == "auto":
        if DUCKDUCKGO_AI_URL:
            return "duckduckgo"
        if os.getenv("OPENROUTER_API_KEY"):
            return "openrouter"
        if os.getenv("DEEPINFRA_API_KEY") or DEEPINFRA_URL:
            return "deepinfra"
        return "pollinations"

    if provider == "openrouter" and not os.getenv("OPENROUTER_API_KEY"):
        _debug("OPENROUTER_API_KEY not set; falling back to pollinations")
        return "pollinations"

    return provider


def _load_dotenv(dotenv_path: Path = DOTENV_PATH) -> None:
    if not dotenv_path.is_file():
        return

    for raw_line in dotenv_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


_load_dotenv()


def _http_json(url: str, *, timeout: int = 12, method: str = "GET", body: Optional[dict[str, Any]] = None, headers: Optional[dict[str, str]] = None) -> Any:
    payload = None
    req_headers = {"User-Agent": "barcode-taric/1.0"}
    if headers:
        req_headers.update(headers)
    if body is not None:
        payload = json.dumps(body).encode("utf-8")
        req_headers["Content-Type"] = "application/json"
    req = urllib.request.Request(url=url, method=method, data=payload, headers=req_headers)
    with urllib.request.urlopen(req, timeout=timeout) as response:
        return json.loads(response.read().decode("utf-8"))


def _http_text(url: str, *, timeout: int = 12, headers: Optional[dict[str, str]] = None) -> str:
    req_headers = {"User-Agent": "barcode-taric/1.0"}
    if headers:
        req_headers.update(headers)

    req = urllib.request.Request(url=url, headers=req_headers)
    with urllib.request.urlopen(req, timeout=timeout) as response:
        return response.read().decode("utf-8", errors="ignore")


def _extract_html_label_pairs(html_text: str) -> dict[str, str]:
    pairs: dict[str, str] = {}
    for label, value in re.findall(r'<td[^>]*class=["\']metadata-label["\'][^>]*>([^<]+)</td>\s*<td>(.*?)</td>', html_text, re.S | re.I):
        pairs[label.strip().lower()] = html.unescape(value.strip())
    return pairs


def _unescape_text(value: str) -> str:
    return html.unescape(re.sub(r"\s+", " ", value.strip()))


def _normalize_text_for_matching(text: str) -> str:
    lowered = text.lower()
    without_diacritics = "".join(
        char for char in unicodedata.normalize("NFKD", lowered) if not unicodedata.combining(char)
    )
    normalized = re.sub(r"[^a-z0-9\s\-]", " ", without_diacritics)
    return re.sub(r"\s+", " ", normalized).strip()


def _normalize_text_for_translation(text: str) -> str:
    lowered = text.lower()
    without_diacritics = "".join(
        char for char in unicodedata.normalize("NFKD", lowered) if not unicodedata.combining(char)
    )
    # Keep Greek and Latin ranges so Greek -> English translations can be applied.
    normalized = re.sub(r"[^a-z0-9\u0370-\u03ff\s\-]", " ", without_diacritics)
    return re.sub(r"\s+", " ", normalized).strip()


def _tokenize_for_matching(text: str) -> list[str]:
    return re.findall(r"[a-z0-9\-]+", text)


def _contains_greek(text: Optional[str]) -> bool:
    return bool(text and re.search(r"[\u0370-\u03FF]", text))


def _contains_latin(text: Optional[str]) -> bool:
    return bool(text and re.search(r"[A-Za-z]", text))


def normalize_to_ean13(code: str) -> Optional[str]:
    digits = re.sub(r"\D", "", code)
    if len(digits) == 13:
        return digits
    if len(digits) == 12:
        checksum = _ean13_checksum(digits)
        return f"{digits}{checksum}"
    if len(digits) == 8:
        return digits.rjust(13, "0")
    return None


def is_valid_ean13(code: str) -> bool:
    if not code.isdigit() or len(code) != 13:
        return False
    return _ean13_checksum(code[:12]) == int(code[-1])


def _ean13_checksum(first12: str) -> int:
    numbers = [int(d) for d in first12]
    total = sum(numbers[i] for i in range(0, 12, 2)) + 3 * sum(numbers[i] for i in range(1, 12, 2))
    return (10 - (total % 10)) % 10


def load_inputs(file_path: Path) -> list[str]:
    suffix = file_path.suffix.lower()
    if suffix == ".txt":
        return [line.strip() for line in file_path.read_text(encoding="utf-8").splitlines() if line.strip()]

    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        values: list[str] = []
        with file_path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.reader(handle, delimiter=delimiter)
            for row in reader:
                for col in row:
                    val = col.strip()
                    if val:
                        values.append(val)
        return values

    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:
            raise ValueError("For Excel input install openpyxl: pip install openpyxl") from exc

        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        excel_values: list[str] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    value = str(cell).strip()
                    if value:
                        excel_values.append(value)
        return excel_values

    raise ValueError("Unsupported input file type. Use .txt, .csv, .tsv, .xlsx")


def fetch_openfoodfacts_product(barcode: str) -> dict[str, Any]:
    try:
        payload = _http_json(OPENFOODFACTS_URL.format(barcode=urllib.parse.quote(barcode)), timeout=10)
    except TimeoutError:
        return {"source": "OpenFoodFacts", "found": False, "error": "timeout"}
    except urllib.error.URLError:
        return {"source": "OpenFoodFacts", "found": False, "error": "network_error"}
    except json.JSONDecodeError:
        return {"source": "OpenFoodFacts", "found": False, "error": "invalid_response"}

    if payload.get("status") != 1:
        return {"source": "OpenFoodFacts", "found": False}

    product = payload.get("product", {})
    return {
        "source": "OpenFoodFacts",
        "found": True,
        "product_name": product.get("product_name"),
        "brand": product.get("brands"),
        "categories": product.get("categories"),
        "description": product.get("generic_name"),
    }


def fetch_wavegrocery_product(barcode: str) -> dict[str, Any]:
    """Fetch product data from public WaveGrocery search endpoint (Thanopoulos default)."""
    url = WAVEGROCERY_PRODUCTS_SEARCH_URL.format(barcode=urllib.parse.quote(barcode))
    try:
        payload = _http_json(url, timeout=12, headers={"appId": "thanopoulos-web"})
    except TimeoutError:
        return {"source": "WaveGrocery", "found": False, "error": "timeout"}
    except urllib.error.URLError:
        return {"source": "WaveGrocery", "found": False, "error": "network_error"}
    except json.JSONDecodeError:
        return {"source": "WaveGrocery", "found": False, "error": "invalid_response"}

    products = (payload or {}).get("data", {}).get("products") or []
    if not products:
        return {"source": "WaveGrocery", "found": False}

    product = products[0]
    product_name = str(product.get("name") or "").strip()
    quantity = str(product.get("quantity") or "").strip()
    description = str(product.get("displayDescription") or "").strip()
    if not description:
        description = " ".join(part for part in (product_name, quantity) if part).strip()

    return {
        "source": "WaveGrocery",
        "found": True,
        "product_name": product_name,
        "brand": product.get("brand"),
        "categories": ", ".join(product.get("collectionTypes") or []) if isinstance(product.get("collectionTypes"), list) else None,
        "description": description,
    }


def fetch_upcitemdb_product(barcode: str) -> dict[str, Any]:
    """Fetch product data from UPC ItemDB (free tier, ~100 req/day, no key)."""
    url = UPCITEMDB_URL.format(barcode=urllib.parse.quote(barcode))
    try:
        payload = _http_json(url, timeout=10)
    except TimeoutError:
        return {"source": "UPCItemDB", "found": False, "error": "timeout"}
    except urllib.error.URLError:
        return {"source": "UPCItemDB", "found": False, "error": "network_error"}
    except json.JSONDecodeError:
        return {"source": "UPCItemDB", "found": False, "error": "invalid_response"}

    items = payload.get("items") or []
    if not items:
        return {"source": "UPCItemDB", "found": False}

    item = items[0]
    return {
        "source": "UPCItemDB",
        "found": True,
        "product_name": item.get("title"),
        "brand": item.get("brand"),
        "categories": item.get("category"),
        "description": item.get("description"),
    }


def fetch_barcode_monster_product(barcode: str) -> dict[str, Any]:
    """Fetch product data from Barcode Monster (free, no key required)."""
    url = BARCODE_MONSTER_URL.format(barcode=urllib.parse.quote(barcode))
    try:
        payload = _http_json(url, timeout=10)
    except TimeoutError:
        return {"source": "BarcodeMonster", "found": False, "error": "timeout"}
    except urllib.error.URLError:
        return {"source": "BarcodeMonster", "found": False, "error": "network_error"}
    except (json.JSONDecodeError, KeyError, TypeError):
        return {"source": "BarcodeMonster", "found": False, "error": "invalid_response"}

    if not payload or not payload.get("product"):
        return {"source": "BarcodeMonster", "found": False}

    return {
        "source": "BarcodeMonster",
        "found": True,
        "product_name": payload.get("product"),
        "brand": payload.get("brand"),
        "categories": payload.get("category"),
        "description": payload.get("description"),
    }


def fetch_go_upc_product(barcode: str) -> dict[str, Any]:
    """Fetch product data from Go-UPC by scraping the search results page."""
    url = GOUNPC_SEARCH_URL.format(barcode=urllib.parse.quote(barcode))
    try:
        html_text = _http_text(
            url,
            timeout=15,
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"},
        )
    except TimeoutError:
        return {"source": "GoUPC", "found": False, "error": "timeout"}
    except urllib.error.URLError:
        return {"source": "GoUPC", "found": False, "error": "network_error"}

    if "no product found" in html_text.lower() or "not found" in html_text.lower():
        return {"source": "GoUPC", "found": False}

    name_match = re.search(r'<h1[^>]*class=["\']product-name["\'][^>]*>(.*?)</h1>', html_text, re.S | re.I)
    product_name = _unescape_text(name_match.group(1)) if name_match else None
    pairs = _extract_html_label_pairs(html_text)
    description_match = re.search(r'<h2>\s*Description\s*</h2>\s*<span>(.*?)</span>', html_text, re.S | re.I)
    description = _unescape_text(description_match.group(1)) if description_match else None

    if not product_name and not pairs:
        return {"source": "GoUPC", "found": False}

    return {
        "source": "GoUPC",
        "found": True,
        "product_name": product_name,
        "brand": pairs.get("brand"),
        "categories": pairs.get("category"),
        "description": description,
    }


def fetch_barcodelookup_product(barcode: str) -> dict[str, Any]:
    """Fetch product data from BarcodeLookup.com by scraping the product page."""
    url = BARCODELOOKUP_URL.format(barcode=urllib.parse.quote(barcode))
    try:
        html_text = _http_text(
            url,
            timeout=15,
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"},
        )
    except TimeoutError:
        return {"source": "BarcodeLookup", "found": False, "error": "timeout"}
    except urllib.error.URLError as exc:
        _debug(f"BarcodeLookup fetch failed: {exc}")
        return {"source": "BarcodeLookup", "found": False, "error": "network_error"}

    if "no product found" in html_text.lower() or "not found" in html_text.lower():
        return {"source": "BarcodeLookup", "found": False}

    title_match = re.search(r'<title>(.*?)</title>', html_text, re.S | re.I)
    product_name = _unescape_text(title_match.group(1)) if title_match else None
    if product_name:
        product_name = re.sub(r"\s*[—-]\s*Barcode Lookup$", "", product_name, flags=re.I).strip()
        if not product_name:
            product_name = None

    pairs = _extract_html_label_pairs(html_text)
    description_match = re.search(r'<h2[^>]*>\s*Description\s*</h2>\s*<div[^>]*>(.*?)</div>', html_text, re.S | re.I)
    description = _unescape_text(description_match.group(1)) if description_match else None

    return {
        "source": "BarcodeLookup",
        "found": bool(product_name or pairs),
        "product_name": product_name,
        "brand": pairs.get("brand"),
        "categories": pairs.get("category"),
        "description": description,
    }


def fetch_ean_search_product(barcode: str) -> dict[str, Any]:
    """Fetch product data from EAN Search by scraping the search results page."""
    url = EAN_SEARCH_URL.format(barcode=urllib.parse.quote(barcode))
    try:
        html_text = _http_text(
            url,
            timeout=15,
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"},
        )
    except TimeoutError:
        return {"source": "EANSearch", "found": False, "error": "timeout"}
    except urllib.error.URLError as exc:
        _debug(f"EANSearch fetch failed: {exc}")
        return {"source": "EANSearch", "found": False, "error": "network_error"}

    lowered = html_text.lower()
    if "no product found" in lowered or "no results found" in lowered or "no results" in lowered:
        return {"source": "EANSearch", "found": False}

    title_match = re.search(r'<title>(.*?)</title>', html_text, re.S | re.I)
    product_name = _unescape_text(title_match.group(1)) if title_match else None
    if product_name:
        product_name = re.sub(r"\s*[—-]\s*EAN Search$", "", product_name, flags=re.I).strip()
        product_name = re.sub(r"\s*[—-]\s*EAN\s*\d+$", "", product_name, flags=re.I).strip()
        if not product_name:
            product_name = None

    pairs = _extract_html_label_pairs(html_text)
    description = None
    meta_desc = re.search(r'<meta\s+name=["\']description["\']\s+content=["\'](.*?)["\']', html_text, re.S | re.I)
    if meta_desc:
        description = _unescape_text(meta_desc.group(1))

    return {
        "source": "EANSearch",
        "found": bool(product_name or pairs or description),
        "product_name": product_name,
        "brand": pairs.get("brand"),
        "categories": pairs.get("category"),
        "description": description,
    }


def fetch_product_multi_source(barcode: str, *, ai_provider: str = "auto") -> dict[str, Any]:
    """Try OpenFoodFacts, UPC ItemDB, Barcode Monster, Go-UPC, EAN Search and BarcodeLookup in order; return first hit."""
    ai_primary = ai_infer_product_from_web(barcode, provider=ai_provider)
    if ai_primary and ai_primary.get("found"):
        _debug(f"Product inferred via AI web primary for {barcode}")
        return ai_primary

    for fetcher in (
        fetch_openfoodfacts_product,
        fetch_upcitemdb_product,
        fetch_barcode_monster_product,
        fetch_go_upc_product,
        fetch_ean_search_product,
        fetch_barcodelookup_product,
    ):
        try:
            result = fetcher(barcode)
            if result.get("found"):
                _debug(f"Product found via {result.get('source')} for {barcode}")
                return result
        except Exception as exc:
            _debug(f"{fetcher.__name__} raised: {exc}")

    return {"source": "none", "found": False}


def parser_rewrite_to_customs_text(text: str) -> str:
    cleaned = _normalize_text_for_translation(text)
    for source, target in sorted(_TRANSLATIONS.items(), key=lambda item: len(item[0]), reverse=True):
        if source in cleaned:
            cleaned = cleaned.replace(source, target)

    tokens = _tokenize_for_matching(cleaned)
    stopwords = {
        "and", "or", "for", "with", "the", "a", "an", "of", "to", "in", "on",
        "των", "και", "σε", "με", "απο", "για", "του", "της", "στο", "στη", "στον",
    }
    filtered = [tok for tok in tokens if tok not in stopwords and len(tok) > 1]

    # Keep insertion order while removing duplicates to stabilize matching.
    unique_tokens = list(dict.fromkeys(filtered))
    return " ".join(unique_tokens)


def _parse_ai_text_response(response: Any) -> Optional[str]:
    if response is None:
        return None
    if isinstance(response, str):
        return response.strip()
    if isinstance(response, dict):
        for key in ("text", "output", "answer", "response", "content"):
            value = response.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()
        if "choices" in response and isinstance(response["choices"], list):
            choice = response["choices"][0]
            if isinstance(choice, dict):
                msg = choice.get("message") or choice
                if isinstance(msg, dict):
                    return _parse_ai_text_response(msg)
                return _parse_ai_text_response(choice)
        if "results" in response and isinstance(response["results"], list):
            return _parse_ai_text_response(response["results"][0])
    return None


def ai_rewrite_to_customs_text(text: str, provider: str = "auto") -> Optional[str]:
    provider = _get_effective_ai_provider(provider)
    prompt = (
        "Rewrite this commercial product text into a concise customs-style description suitable for EU TARIC classification. "
        "Focus on material composition, product use/function, and any relevant customs classification hints. "
        "For beverages, include type. For tobacco/vaping products, include nicotine status. "
        "Return only the short classification description in English with no extra commentary. "
        f"Product: {text}"
    )

    if provider == "none":
        return None

    try:
        if provider == "pollinations":
            encoded = urllib.parse.quote(prompt)
            response = _http_text(POLLINATIONS_URL.format(prompt=encoded), timeout=18).strip()
            return response[:500] if response else None

        if provider == "openrouter":
            api_key = os.getenv("OPENROUTER_API_KEY")
            if not api_key:
                _debug("OpenRouter key missing; falling back to pollinations")
                return ai_rewrite_to_customs_text(text, provider="pollinations")

            payload = {
                "model": "meta-llama/llama-3.3-70b-instruct:free",
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0,
            }
            response = _http_json(
                OPENROUTER_URL,
                method="POST",
                body=payload,
                headers={"Authorization": f"Bearer {api_key}"},
                timeout=20,
            )
            content = _parse_ai_text_response(response)
            if content:
                return content[:500]
            _debug("OpenRouter response missing content; falling back to pollinations")
            return ai_rewrite_to_customs_text(text, provider="pollinations")

        if provider == "deepinfra":
            if not DEEPINFRA_URL:
                _debug("DeepInfra URL not configured; falling back to pollinations")
                return ai_rewrite_to_customs_text(text, provider="pollinations")
            payload = {
                "model": DEEPINFRA_MODEL,
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0,
                "max_tokens": 150,
            }
            headers = {"Content-Type": "application/json"}
            if os.getenv("DEEPINFRA_API_KEY"):
                headers["Authorization"] = f"Bearer {os.getenv('DEEPINFRA_API_KEY')}"
            response = _http_json(DEEPINFRA_URL, method="POST", body=payload, headers=headers, timeout=25)
            content = _parse_ai_text_response(response)
            if content:
                return content[:500]
            _debug("DeepInfra response missing content; falling back to pollinations")
            return ai_rewrite_to_customs_text(text, provider="pollinations")

        if provider == "duckduckgo":
            if not DUCKDUCKGO_AI_URL:
                _debug("DuckDuckGo AI URL not configured; falling back to pollinations")
                return ai_rewrite_to_customs_text(text, provider="pollinations")
            payload = {"question": prompt}
            response = _http_json(DUCKDUCKGO_AI_URL, method="POST", body=payload, headers={"Content-Type": "application/json"}, timeout=25)
            content = _parse_ai_text_response(response)
            if content:
                return content[:500]
            _debug("DuckDuckGo AI response missing content; falling back to pollinations")
            return ai_rewrite_to_customs_text(text, provider="pollinations")

        if provider == "nerve":
            if not NERVE_URL:
                _debug("Nerve URL not configured; falling back to pollinations")
                return ai_rewrite_to_customs_text(text, provider="pollinations")
            payload = {"prompt": prompt, "max_new_tokens": 150}
            response = _http_json(NERVE_URL, method="POST", body=payload, headers={"Content-Type": "application/json"}, timeout=25)
            content = _parse_ai_text_response(response)
            if content:
                return content[:500]
            _debug("Nerve/Text-Generation response missing content; falling back to pollinations")
            return ai_rewrite_to_customs_text(text, provider="pollinations")

    except (urllib.error.URLError, TimeoutError, KeyError, IndexError, TypeError, json.JSONDecodeError) as exc:
        _debug(f"AI rewrite failed ({provider}): {exc}")
        if provider != "pollinations":
            return ai_rewrite_to_customs_text(text, provider="pollinations")
        return None

    return None


def ai_translate_text(text: str, *, target_language: str, provider: str = "auto") -> Optional[str]:
    provider = _get_effective_ai_provider(provider)
    if provider == "none" or not text.strip():
        return None

    language_label = "Greek" if target_language.lower().startswith("el") else "English"
    prompt = (
        f"Translate the following product text to {language_label}. "
        "Keep product names and brands unchanged when appropriate. "
        "Return only the translated text with no extra commentary. "
        f"Text: {text}"
    )

    try:
        if provider == "pollinations":
            encoded = urllib.parse.quote(prompt)
            response = _http_text(POLLINATIONS_URL.format(prompt=encoded), timeout=18).strip()
            return response[:450] if response else None

        if provider == "openrouter":
            api_key = os.getenv("OPENROUTER_API_KEY")
            if not api_key:
                return None
            payload = {
                "model": "meta-llama/llama-3.3-70b-instruct:free",
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0,
            }
            response = _http_json(
                OPENROUTER_URL,
                method="POST",
                body=payload,
                headers={"Authorization": f"Bearer {api_key}"},
                timeout=20,
            )
            content = _parse_ai_text_response(response)
            return content[:450] if content else None

        if provider == "deepinfra":
            if not DEEPINFRA_URL:
                return None
            payload = {
                "model": DEEPINFRA_MODEL,
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0,
                "max_tokens": 220,
            }
            headers = {"Content-Type": "application/json"}
            if os.getenv("DEEPINFRA_API_KEY"):
                headers["Authorization"] = f"Bearer {os.getenv('DEEPINFRA_API_KEY')}"
            response = _http_json(DEEPINFRA_URL, method="POST", body=payload, headers=headers, timeout=25)
            content = _parse_ai_text_response(response)
            return content[:450] if content else None

        if provider == "duckduckgo":
            if not DUCKDUCKGO_AI_URL:
                return None
            response = _http_json(
                DUCKDUCKGO_AI_URL,
                method="POST",
                body={"question": prompt},
                headers={"Content-Type": "application/json"},
                timeout=25,
            )
            content = _parse_ai_text_response(response)
            return content[:450] if content else None

        if provider == "nerve":
            if not NERVE_URL:
                return None
            response = _http_json(
                NERVE_URL,
                method="POST",
                body={"prompt": prompt, "max_new_tokens": 220},
                headers={"Content-Type": "application/json"},
                timeout=25,
            )
            content = _parse_ai_text_response(response)
            return content[:450] if content else None
    except (urllib.error.URLError, TimeoutError, KeyError, IndexError, TypeError, json.JSONDecodeError) as exc:
        _debug(f"AI translation failed ({provider}): {exc}")
        return None

    return None


def _duckduckgo_web_search_context(query: str, *, limit: int = 5) -> str:
    search_url = "https://html.duckduckgo.com/html/?q={query}".format(query=urllib.parse.quote(query))
    try:
        html_text = _http_text(
            search_url,
            timeout=15,
            headers={"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/120 Safari/537.36"},
        )
    except (urllib.error.URLError, TimeoutError):
        return ""

    items: list[str] = []
    result_blocks = re.findall(r'<div[^>]*class=["\']result["\'][^>]*>(.*?)</div>\s*</div>', html_text, re.S | re.I)
    for block in result_blocks[:limit]:
        title_match = re.search(r'class=["\']result__a["\'][^>]*>(.*?)</a>', block, re.S | re.I)
        snippet_match = re.search(r'class=["\']result__snippet["\'][^>]*>(.*?)</a>|class=["\']result__snippet["\'][^>]*>(.*?)</div>', block, re.S | re.I)
        title = _unescape_text(re.sub(r"<[^>]+>", " ", title_match.group(1))) if title_match else ""
        snippet_raw = ""
        if snippet_match:
            snippet_raw = snippet_match.group(1) or snippet_match.group(2) or ""
        snippet = _unescape_text(re.sub(r"<[^>]+>", " ", snippet_raw)) if snippet_raw else ""
        line = " - ".join(part for part in (title, snippet) if part)
        if line:
            items.append(line)

    return "\n".join(items)


def _looks_low_quality_product_name(name: str, *, description: str = "", brand: str = "") -> bool:
    normalized = _normalize_text_for_matching(name)
    if not normalized:
        return True

    tokens = [tok for tok in re.split(r"\s+", normalized) if tok]
    if len(tokens) >= 14:
        return True

    spam_tokens = {
        "women", "woman", "men", "man", "dark", "circle", "blemishes", "cover", "kit", "set",
        "skin", "care", "cosmetics", "makeup", "health", "beauty", "personal",
    }
    spam_hits = sum(1 for tok in tokens if tok in spam_tokens)
    if spam_hits >= 5:
        return True

    if description:
        norm_desc = _normalize_text_for_matching(description)
        if norm_desc and normalized == norm_desc:
            return True

    if brand:
        norm_brand = _normalize_text_for_matching(brand)
        if norm_brand and norm_brand in normalized and len(tokens) > 10:
            return True

    return False


def _derive_fallback_product_name(*, brand: str = "", description: str = "", categories: str = "") -> str:
    descriptor = ""
    if description:
        blocked = {
            "women", "woman", "men", "man", "dark", "circle", "blemishes", "cover", "kit", "set",
            "skin", "care", "cosmetics", "makeup", "health", "beauty", "personal", "for", "with", "and",
        }
        raw_tokens = re.findall(r"[A-Za-z0-9\u0370-\u03FF\u1F00-\u1FFF]+", description)
        keep_tokens = [tok for tok in raw_tokens if _normalize_text_for_matching(tok) not in blocked]
        descriptor = " ".join(keep_tokens[:3]).strip()

    if not descriptor and categories:
        leaf = categories.split(">")[-1].strip()
        descriptor = leaf

    if brand and descriptor:
        if _normalize_text_for_matching(brand) in _normalize_text_for_matching(descriptor):
            return descriptor
        return f"{brand} {descriptor}".strip()
    return (brand or descriptor or "").strip()


def _sanitize_product_name(name: str, *, brand: str = "", description: str = "", categories: str = "") -> str:
    clean_name = str(name or "").strip()
    if not clean_name:
        return ""
    if not _looks_low_quality_product_name(clean_name, description=description, brand=brand):
        return clean_name
    fallback = _derive_fallback_product_name(brand=brand, description=description, categories=categories)
    return fallback or ""


def ai_infer_product_from_web(barcode: str, provider: str = "auto") -> Optional[dict[str, Any]]:
    provider = _get_effective_ai_provider(provider)
    if provider == "none":
        return None

    web_context = _duckduckgo_web_search_context(barcode, limit=6)
    if web_context:
        prompt = (
            "You are given web search snippets for a barcode. Infer likely product metadata. "
            "Return ONLY valid JSON with keys: product_name, brand, categories, description, confidence. "
            "Description should be in English and concise. If uncertain, keep fields empty strings and confidence='low'. "
            f"Barcode: {barcode}\n"
            f"Web snippets:\n{web_context}"
        )
    else:
        prompt = (
            "Infer likely product metadata for this barcode. "
            "If your runtime supports web search/browsing, use it before answering. "
            "Return ONLY valid JSON with keys: product_name, brand, categories, description, confidence. "
            "If uncertain, keep fields empty strings and confidence='low'. "
            f"Barcode: {barcode}"
        )

    try:
        if provider == "pollinations":
            encoded = urllib.parse.quote(prompt)
            response = _http_text(POLLINATIONS_URL.format(prompt=encoded), timeout=20).strip()
        elif provider == "openrouter":
            api_key = os.getenv("OPENROUTER_API_KEY")
            if not api_key:
                return None
            payload = {
                "model": "meta-llama/llama-3.3-70b-instruct:free",
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.1,
            }
            response = _http_json(
                OPENROUTER_URL,
                method="POST",
                body=payload,
                headers={"Authorization": f"Bearer {api_key}"},
                timeout=20,
            )
            response = _parse_ai_text_response(response)
        elif provider == "deepinfra":
            if not DEEPINFRA_URL:
                return None
            payload = {
                "model": DEEPINFRA_MODEL,
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.1,
                "max_tokens": 260,
            }
            headers = {"Content-Type": "application/json"}
            if os.getenv("DEEPINFRA_API_KEY"):
                headers["Authorization"] = f"Bearer {os.getenv('DEEPINFRA_API_KEY')}"
            response = _http_json(DEEPINFRA_URL, method="POST", body=payload, headers=headers, timeout=25)
            response = _parse_ai_text_response(response)
        elif provider == "duckduckgo":
            if not DUCKDUCKGO_AI_URL:
                return None
            response = _http_json(
                DUCKDUCKGO_AI_URL,
                method="POST",
                body={"question": prompt},
                headers={"Content-Type": "application/json"},
                timeout=25,
            )
            response = _parse_ai_text_response(response)
        elif provider == "nerve":
            if not NERVE_URL:
                return None
            response = _http_json(
                NERVE_URL,
                method="POST",
                body={"prompt": prompt, "max_new_tokens": 260},
                headers={"Content-Type": "application/json"},
                timeout=25,
            )
            response = _parse_ai_text_response(response)
        else:
            return None
    except (urllib.error.URLError, TimeoutError, KeyError, IndexError, TypeError, json.JSONDecodeError) as exc:
        _debug(f"AI web infer failed ({provider}): {exc}")
        return None

    if not response:
        return None

    raw_text = response.strip()
    raw_text = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw_text, flags=re.I | re.S)
    try:
        parsed = json.loads(raw_text)
    except json.JSONDecodeError:
        return None

    if not isinstance(parsed, dict):
        return None

    name = str(parsed.get("product_name") or "").strip()
    desc = str(parsed.get("description") or "").strip()
    brand = str(parsed.get("brand") or "").strip()
    categories = str(parsed.get("categories") or "").strip()
    confidence = str(parsed.get("confidence") or "").strip().lower()
    if not name and not desc:
        return None
    if confidence in {"low", "very_low", "unknown"} and not web_context:
        return None

    if name and _looks_low_quality_product_name(name, description=desc, brand=brand):
        _debug(f"Discarding low-quality AI inferred product name for {barcode}: {name!r}")
        name = ""

    if not name and not desc:
        return None

    return {
        "source": "AIWebSearch",
        "found": True,
        "product_name": name,
        "brand": brand or None,
        "categories": categories or None,
        "description": desc,
        "confidence": str(parsed.get("confidence") or "").strip() or None,
        "web_context_used": bool(web_context),
    }


def _is_placeholder_description(value: Optional[str]) -> bool:
    if not value:
        return True
    normalized = _normalize_text_for_matching(value)
    placeholder_patterns = {
        "no description",
        "description not found",
        "not found",
        "unknown",
        "n a",
        "na",
    }
    if len(normalized) < 12:
        return True
    return any(pattern in normalized for pattern in placeholder_patterns)


def ai_enrich_product_description_greek(
    *,
    product_name: str,
    brand: Optional[str],
    categories: Optional[str],
    barcode: Optional[str],
    provider: str = "auto",
) -> Optional[str]:
    provider = _get_effective_ai_provider(provider)
    if provider == "none":
        return None

    prompt = (
        "Δημιούργησε μία φυσική εμπορική περιγραφή στα ελληνικά για προϊόν λιανικής. "
        "Χρησιμοποίησε τα διαθέσιμα στοιχεία (όνομα, μάρκα, κατηγορία, barcode). "
        "Αν φαίνεται ότι είναι ποτό, πρόσθεσε πιθανό τύπο και συσκευασία (π.χ. 750ml) μόνο αν είναι λογικό. "
        "Δώσε 1 πρόταση, χωρίς markdown, χωρίς bullets, χωρίς αποποίηση ευθύνης. "
        f"Όνομα: {product_name or ''} | Μάρκα: {brand or ''} | Κατηγορία: {categories or ''} | Barcode: {barcode or ''}"
    )

    try:
        if provider == "pollinations":
            encoded = urllib.parse.quote(prompt)
            response = _http_text(POLLINATIONS_URL.format(prompt=encoded), timeout=18).strip()
            return response[:350] if response else None

        if provider == "openrouter":
            api_key = os.getenv("OPENROUTER_API_KEY")
            if not api_key:
                return None
            payload = {
                "model": "meta-llama/llama-3.3-70b-instruct:free",
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.2,
            }
            response = _http_json(
                OPENROUTER_URL,
                method="POST",
                body=payload,
                headers={"Authorization": f"Bearer {api_key}"},
                timeout=20,
            )
            content = _parse_ai_text_response(response)
            return content[:350] if content else None

        if provider == "deepinfra":
            if not DEEPINFRA_URL:
                return None
            payload = {
                "model": DEEPINFRA_MODEL,
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.2,
                "max_tokens": 180,
            }
            headers = {"Content-Type": "application/json"}
            if os.getenv("DEEPINFRA_API_KEY"):
                headers["Authorization"] = f"Bearer {os.getenv('DEEPINFRA_API_KEY')}"
            response = _http_json(DEEPINFRA_URL, method="POST", body=payload, headers=headers, timeout=25)
            content = _parse_ai_text_response(response)
            return content[:350] if content else None

        if provider == "duckduckgo":
            if not DUCKDUCKGO_AI_URL:
                return None
            response = _http_json(
                DUCKDUCKGO_AI_URL,
                method="POST",
                body={"question": prompt},
                headers={"Content-Type": "application/json"},
                timeout=25,
            )
            content = _parse_ai_text_response(response)
            return content[:350] if content else None

        if provider == "nerve":
            if not NERVE_URL:
                return None
            response = _http_json(
                NERVE_URL,
                method="POST",
                body={"prompt": prompt, "max_new_tokens": 180},
                headers={"Content-Type": "application/json"},
                timeout=25,
            )
            content = _parse_ai_text_response(response)
            return content[:350] if content else None
    except (urllib.error.URLError, TimeoutError, KeyError, IndexError, TypeError, json.JSONDecodeError) as exc:
        _debug(f"AI Greek description enrichment failed ({provider}): {exc}")
        return None

    return None


def ai_inspect_and_correct_customs_text(
    match: Optional[TaricEntry],
    customs_text: str,
    commercial_text: str,
    provider: str = "auto",
) -> Optional[str]:
    if provider == "none":
        return None

    match_label = (
        f"{match.taric_code} ({match.description})" if match else "no current TARIC match"
    )

    prompt = (
        "Inspect the previously selected EU TARIC classification and correct it if needed. "
        "Commercial product text: {commercial_text} "
        "Current customs-style description: {customs_text} "
        "Current TARIC match: {match_label} "
        "If the match is incorrect, rewrite a concise customs-style description in English that better reflects the product and can be used for EU TARIC matching. "
        "If the match is already correct, return the same description or a slightly improved version. "
        "Do not include any extra commentary."
    ).format(
        commercial_text=commercial_text,
        customs_text=customs_text,
        match_label=match_label,
    )

    return ai_rewrite_to_customs_text(prompt, provider=provider)


def search_products_by_description(search_text: str, ai_provider: str = "auto") -> list[dict[str, Any]]:
    """Search database and web for products matching description, retrieve barcodes and TARIC codes."""
    if not HAS_DATABASE:
        _debug("Database not available for description search")
        return []
    
    # First check local database for existing matches
    db_results = search_by_description(search_text, limit=20)
    results = []
    
    for record in db_results:
        results.append({
            "input": search_text,
            "barcode": record.barcode,
            "product_name": record.product_name,
            "product_description": record.description,
            "commercial_text": record.commercial_text,
            "customs_text": record.customs_text,
            "match": {
                "taric_code": record.taric_code,
                "hs4": record.hs4,
                "description": record.taric_description,
            },
            "source": "database"
        })
    
    return results


def _score_taric_entry(normalized_text: str, query_tokens: set[str], entry: TaricEntry) -> int:
    text = f" {normalized_text} "
    score = 0
    normalized_description = _normalize_text_for_matching(entry.description)

    for keyword in entry.keywords:
        normalized_keyword = _normalize_text_for_matching(keyword)
        keyword_tokens = _tokenize_for_matching(normalized_keyword)
        if not keyword_tokens:
            continue

        if f" {normalized_keyword} " in text:
            score += 4 + len(keyword_tokens)
        else:
            overlap = len(query_tokens.intersection(keyword_tokens))
            score += overlap

    vaping_keywords = {"vape", "eliquid", "ecig", "ecigarette", "puff", "inhalation", "nicotine", "nic salt"}
    has_vaping_context = bool(query_tokens.intersection(vaping_keywords))

    food_beverage_keywords = {"juice", "drink", "beverage", "flavor", "flavour", "fruit", "pear", "apple", "orange", "grape", "concentrate", "syrup", "aroma"}
    has_food_context = bool(query_tokens.intersection(food_beverage_keywords))

    tissue_keywords = {"toilet", "paper", "tissue", "bathroom", "ply", "roll"}
    has_tissue_context = bool(query_tokens.intersection(tissue_keywords))

    cosmetics_keywords = {
        "cosmetic", "cosmetics", "makeup", "concealer", "mascara", "lipstick",
        "foundation", "eyeliner", "skincare", "beauty",
    }
    has_cosmetics_context = bool(query_tokens.intersection(cosmetics_keywords))

    if has_food_context and not has_vaping_context:
        if entry.hs4.startswith("2404"):
            score -= 10

    if has_vaping_context and entry.hs4.startswith("2404"):
        score += 5

    if has_tissue_context:
        if entry.hs4 == "4818":
            score += 4
        if entry.hs4 == "4809":
            score -= 2

    if has_cosmetics_context:
        if entry.hs4 == "3304":
            score += 6
        if entry.hs4 == "9619":
            score -= 4
        if entry.hs4.startswith("22"):
            score -= 8

    if "nicotine" in query_tokens and has_vaping_context:
        if "with nicotine" in normalized_description:
            score += 5
        if "without nicotine" in normalized_description:
            score -= 2

    return score


def _taric_entry_by_code(taric_code: str, catalog: Iterable[TaricEntry] = DEFAULT_TARIC_CATALOG) -> Optional[TaricEntry]:
    cleaned_code = re.sub(r"\D", "", taric_code)
    for entry in catalog:
        if entry.taric_code == cleaned_code:
            return entry
    return None


def _top_taric_candidates(customs_text: str, catalog: Iterable[TaricEntry] = DEFAULT_TARIC_CATALOG, limit: int = 5) -> list[dict[str, str]]:
    normalized_text = _normalize_text_for_matching(customs_text)
    query_tokens = set(_tokenize_for_matching(normalized_text))
    ranked: list[tuple[int, TaricEntry]] = []

    for entry in catalog:
        ranked.append((_score_taric_entry(normalized_text, query_tokens, entry), entry))

    ranked.sort(key=lambda item: item[0], reverse=True)
    candidates: list[dict[str, str]] = []
    for score, entry in ranked[:limit]:
        if score <= 0:
            continue
        candidates.append({
            "taric_code": entry.taric_code,
            "hs4": entry.hs4,
            "description": entry.description,
        })
    return candidates


def best_taric_match(customs_text: str, catalog: Iterable[TaricEntry] = DEFAULT_TARIC_CATALOG) -> Optional[TaricEntry]:
    normalized_text = _normalize_text_for_matching(customs_text)
    query_tokens = set(_tokenize_for_matching(normalized_text))

    best: tuple[int, Optional[TaricEntry]] = (0, None)
    for entry in catalog:
        score = _score_taric_entry(normalized_text, query_tokens, entry)
        if score > best[0]:
            best = (score, entry)

    if best[0] < 3:
        return None
    return best[1]


def ai_validate_taric_match(
    match: Optional[TaricEntry],
    customs_text: str,
    commercial_text: str,
    provider: str = "auto",
    catalog: Iterable[TaricEntry] = DEFAULT_TARIC_CATALOG,
) -> Optional[dict[str, Any]]:
    provider = _get_effective_ai_provider(provider)
    if provider == "none":
        return None

    match_label = f"{match.taric_code} ({match.description})" if match else "no current TARIC match"
    candidates = _top_taric_candidates(customs_text, catalog=catalog, limit=5)
    prompt = (
        "Validate the EU TARIC classification using the latest available catalog context. "
        "Use the commercial text, the customs-style description, the current TARIC match, and the candidate TARIC codes. "
        "Decide whether the current match is correct. If it is wrong, choose the best candidate or return a corrected customs description that leads to a better match. "
        "Return ONLY valid JSON with these keys: approved (boolean), taric_code (string or null), hs4 (string or null), customs_text (string or null), reason (string), confidence (string). "
        "Commercial product text: {commercial_text} "
        "Current customs-style description: {customs_text} "
        "Current TARIC match: {match_label} "
        "Candidate TARIC codes: {candidates}"
    ).format(
        commercial_text=commercial_text,
        customs_text=customs_text,
        match_label=match_label,
        candidates=json.dumps(candidates, ensure_ascii=False),
    )

    try:
        if provider == "pollinations":
            encoded = urllib.parse.quote(prompt)
            response = _http_text(POLLINATIONS_URL.format(prompt=encoded), timeout=18).strip()
        elif provider == "openrouter":
            api_key = os.getenv("OPENROUTER_API_KEY")
            if not api_key:
                return None
            payload = {
                "model": "meta-llama/llama-3.3-70b-instruct:free",
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0,
            }
            response = _http_json(
                OPENROUTER_URL,
                method="POST",
                body=payload,
                headers={"Authorization": f"Bearer {api_key}"},
                timeout=20,
            )
            response = _parse_ai_text_response(response)
        elif provider == "deepinfra":
            if not DEEPINFRA_URL:
                return None
            payload = {
                "model": DEEPINFRA_MODEL,
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0,
                "max_tokens": 220,
            }
            headers = {"Content-Type": "application/json"}
            if os.getenv("DEEPINFRA_API_KEY"):
                headers["Authorization"] = f"Bearer {os.getenv('DEEPINFRA_API_KEY')}"
            response = _http_json(DEEPINFRA_URL, method="POST", body=payload, headers=headers, timeout=25)
            response = _parse_ai_text_response(response)
        elif provider == "duckduckgo":
            if not DUCKDUCKGO_AI_URL:
                return None
            payload = {"question": prompt}
            response = _http_json(DUCKDUCKGO_AI_URL, method="POST", body=payload, headers={"Content-Type": "application/json"}, timeout=25)
            response = _parse_ai_text_response(response)
        elif provider == "nerve":
            if not NERVE_URL:
                return None
            payload = {"prompt": prompt, "max_new_tokens": 220}
            response = _http_json(NERVE_URL, method="POST", body=payload, headers={"Content-Type": "application/json"}, timeout=25)
            response = _parse_ai_text_response(response)
        else:
            return None
    except (urllib.error.URLError, TimeoutError, KeyError, IndexError, TypeError, json.JSONDecodeError) as exc:
        _debug(f"AI TARIC validation failed ({provider}): {exc}")
        return None

    if not response:
        return None

    raw_text = response.strip()
    raw_text = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw_text, flags=re.I | re.S)
    try:
        parsed = json.loads(raw_text)
    except json.JSONDecodeError:
        return None

    if not isinstance(parsed, dict):
        return None

    parsed["approved"] = bool(parsed.get("approved"))
    if parsed.get("taric_code"):
        parsed["taric_code"] = re.sub(r"\D", "", str(parsed["taric_code"]))
    if parsed.get("hs4"):
        parsed["hs4"] = re.sub(r"\D", "", str(parsed["hs4"]))
    return parsed


def resolve_item(item: str, *, ai_provider: str = "auto", store_in_db: bool = True) -> dict[str, Any]:
    query = item.strip()
    normalized_barcode = normalize_to_ean13(query)
    effective_provider = _get_effective_ai_provider(ai_provider)

    source_data: dict[str, Any] = {}
    commercial_text = query
    product_name = ""
    product_name_en = ""
    product_description = ""
    description_en = ""
    
    if normalized_barcode:
        source_data = fetch_product_multi_source(normalized_barcode, ai_provider=effective_provider)
        if source_data.get("found"):
            product_name = source_data.get("product_name", "")
            product_description = source_data.get("description", "")
            product_name = _sanitize_product_name(
                product_name,
                brand=str(source_data.get("brand") or ""),
                description=str(product_description or product_name or ""),
                categories=str(source_data.get("categories") or ""),
            )
            source_data["product_name"] = product_name

    if product_name and effective_provider != "none":
        if _contains_greek(product_name):
            translated_name = ai_translate_text(product_name, target_language="en", provider=effective_provider)
            if translated_name and not _contains_greek(translated_name):
                product_name_en = translated_name
            else:
                product_name_en = parser_rewrite_to_customs_text(product_name)
        else:
            product_name_en = product_name
    else:
        product_name_en = product_name

    if normalized_barcode and effective_provider != "none":
        # Keep DB columns semantically stable: description=Greek, commercial_text=English.
        if product_description:
            if _contains_greek(product_description):
                translated_en = ai_translate_text(product_description, target_language="en", provider=effective_provider)
                if translated_en and not _contains_greek(translated_en):
                    description_en = translated_en
                else:
                    description_en = parser_rewrite_to_customs_text(product_description)
            elif _contains_latin(product_description) and not _contains_greek(product_description):
                description_en = product_description

        if description_en and not _contains_greek(product_description):
            translated_el = ai_translate_text(description_en, target_language="el", provider=effective_provider)
            if translated_el:
                product_description = translated_el

        if product_description and _contains_greek(product_description) and _contains_greek(description_en):
            translated_en = ai_translate_text(product_description, target_language="en", provider=effective_provider)
            if translated_en:
                description_en = translated_en

    if normalized_barcode and _is_placeholder_description(product_description) and effective_provider != "none":
        enriched_description = ai_enrich_product_description_greek(
            product_name=product_name,
            brand=source_data.get("brand"),
            categories=source_data.get("categories"),
            barcode=normalized_barcode,
            provider=effective_provider,
        )
        if enriched_description:
            product_description = enriched_description
            if source_data:
                source_data["description"] = enriched_description

    if normalized_barcode and not source_data.get("found"):
        # Avoid false TARIC matches from bare barcode digits when we have no evidence.
        return {
            "input": query,
            "barcode": normalized_barcode,
            "valid_ean13": is_valid_ean13(normalized_barcode),
            "source": {"source": "none", "found": False},
            "product_name": "",
            "product_name_en": "",
            "product_description": "",
            "commercial_text": "",
            "customs_text": "",
            "ai_inspection": {
                "provider": effective_provider,
                "corrected_customs_text": None,
            },
            "ai_validation": {"checked": False, "approved": None},
            "match": {
                "taric_code": None,
                "hs4": None,
                "description": "No confident match",
            },
            "validation": {
                "checked": False,
                "stored": False,
            },
        }

    if not description_en and product_description:
        if _contains_greek(product_description) and effective_provider != "none":
            translated_en = ai_translate_text(product_description, target_language="en", provider=effective_provider)
            if translated_en and not _contains_greek(translated_en):
                description_en = translated_en
            else:
                description_en = parser_rewrite_to_customs_text(product_description)
        else:
            description_en = product_description

    commercial_parts = [
        product_name_en or product_name,
        source_data.get("categories"),
        source_data.get("brand"),
        description_en,
    ]
    commercial_text = " ".join(str(p) for p in commercial_parts if p and str(p) not in ("None", "")) or query

    parser_text = parser_rewrite_to_customs_text(commercial_text)
    ai_text = ai_rewrite_to_customs_text(commercial_text, provider=effective_provider)
    customs_text = parser_rewrite_to_customs_text(ai_text) if ai_text else parser_text

    corrected_customs_text = None
    match = best_taric_match(customs_text)
    validation_result: Optional[dict[str, Any]] = None

    if effective_provider != "none":
        verified_text = ai_inspect_and_correct_customs_text(
            match,
            customs_text,
            commercial_text,
            provider=effective_provider,
        )
        if verified_text:
            corrected_customs_text = parser_rewrite_to_customs_text(verified_text)
            if corrected_customs_text != customs_text:
                _debug("AI inspection revised customs text; re-running TARIC matching")
                customs_text = corrected_customs_text
                match = best_taric_match(customs_text)

        validation_result = ai_validate_taric_match(
            match,
            customs_text,
            commercial_text,
            provider=effective_provider,
        )

        if validation_result:
            validated_customs_text = validation_result.get("customs_text")
            if isinstance(validated_customs_text, str) and validated_customs_text.strip():
                corrected_customs_text = parser_rewrite_to_customs_text(validated_customs_text)
                if corrected_customs_text:
                    customs_text = corrected_customs_text
                    match = best_taric_match(customs_text)

            validated_code = validation_result.get("taric_code")
            validated_entry = _taric_entry_by_code(str(validated_code)) if validated_code else None
            if validated_entry:
                match = validated_entry
            elif match is None:
                match = best_taric_match(customs_text)

            if validation_result.get("approved") is False:
                approved_code = validation_result.get("taric_code")
                if approved_code and not validated_entry:
                    fallback_text = validation_result.get("customs_text")
                    if isinstance(fallback_text, str) and fallback_text.strip():
                        customs_text = parser_rewrite_to_customs_text(fallback_text)
                        match = best_taric_match(customs_text)

    result = {
        "input": query,
        "barcode": normalized_barcode,
        "valid_ean13": is_valid_ean13(normalized_barcode) if normalized_barcode else False,
        "source": source_data or {"source": "direct_input"},
        "product_name": product_name,
        "product_name_en": product_name_en,
        "product_description": product_description,
        "commercial_text": commercial_text,
        "customs_text": customs_text,
        "ai_inspection": {
            "provider": effective_provider,
            "corrected_customs_text": corrected_customs_text,
        },
        "ai_validation": validation_result or {"checked": False, "approved": None},
        "match": {
            "taric_code": match.taric_code if match else None,
            "hs4": match.hs4 if match else None,
            "description": match.description if match else "No confident match",
        },
        "validation": {
            "checked": True,
            "stored": bool(match),
        },
    }
    
    # Store only after a successful TARIC check.
    if store_in_db and HAS_DATABASE and match:
        product = ProductRecord(
            barcode=normalized_barcode,
            product_name=product_name,
            product_name_en=product_name_en,
            description=product_description,
            commercial_text=commercial_text,
            customs_text=customs_text,
            taric_code=match.taric_code if match else None,
            hs4=match.hs4 if match else None,
            taric_description=match.description if match else "No confident match",
            source=source_data.get("source", "api")
        )
        add_product(product)
    
    return result


def parse_args(argv: Optional[list[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Match barcode or product description to TARIC code (bidirectional search)")
    parser.add_argument("query", nargs="?", help="Single barcode or product description")
    parser.add_argument("--file", type=Path, help="Batch file with inputs (.txt/.csv/.xlsx)")
    parser.add_argument(
        "--ai-provider",
        choices=["auto", "pollinations", "openrouter", "deepinfra", "duckduckgo", "nerve", "none"],
        default="auto",
    )
    parser.add_argument("--output", type=Path, default=Path("taric_output.json"), help="Output JSON path")
    parser.add_argument("--mode", choices=["barcode", "description", "auto"], default="auto",
                       help="Search mode: 'barcode' (13-digit search), 'description' (FTS), 'auto' (detect)")
    parser.add_argument("--no-db", action="store_true", help="Disable database storage")
    parser.add_argument("--export-db", action="store_true", help="Export database to JSON")
    parser.add_argument("--import-db", type=Path, help="Import products from JSON")
    parser.add_argument("--list-db", action="store_true", help="List all products in database (detailed view)")
    parser.add_argument("--view-db", action="store_true", help="Alias for --list-db (view database)")
    return parser.parse_args(argv)


def main(argv: Optional[list[str]] = None) -> int:
    args = parse_args(argv)
    
    # Handle database export/import/view operations
    if args.export_db:
        if HAS_DATABASE:
            export_database_to_json()
            return 0
        else:
            print("Database module not available", file=sys.stderr)
            return 1
    
    if args.import_db:
        if HAS_DATABASE and args.import_db.exists():
            import_database_from_json(args.import_db)
            return 0
        else:
            print(f"Cannot import: file not found or database unavailable", file=sys.stderr)
            return 1
    
    # List/view database contents
    if args.list_db or args.view_db:
        if HAS_DATABASE:
            from product_database import list_products_detailed
            list_products_detailed()
            return 0
        else:
            print("Database module not available", file=sys.stderr)
            return 1
    
    # Initialize database if enabled
    if HAS_DATABASE and not args.no_db:
        init_database()

    items: list[str] = []
    if args.query:
        items.append(args.query)
    if args.file:
        try:
            items.extend(load_inputs(args.file))
        except (ValueError, OSError) as exc:
            print(f"Failed to read {args.file}: {exc}", file=sys.stderr)
            return 2

    if not items:
        print("Provide a query or --file", file=sys.stderr)
        return 2

    results = []
    for item in items:
        # Auto-detect search mode or use specified mode
        mode = args.mode
        if mode == "auto":
            normalized = normalize_to_ean13(item)
            mode = "barcode" if normalized else "description"
        
        if mode == "barcode":
            # Standard barcode lookup
            result = resolve_item(item, ai_provider=args.ai_provider, store_in_db=not args.no_db)
            results.append(result)
        else:  # mode == "description"
            # Description-based search
            db_results = search_products_by_description(item, ai_provider=args.ai_provider)
            if db_results:
                results.extend(db_results)
            else:
                # If no database results, try as direct input for TARIC matching
                result = resolve_item(item, ai_provider=args.ai_provider, store_in_db=not args.no_db)
                results.append(result)
    
    args.output.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(results, ensure_ascii=False, indent=2))
    print(f"Saved {len(results)} result(s) to {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
