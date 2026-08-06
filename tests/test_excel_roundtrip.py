from pathlib import Path

from barcodetaric import repo
from barcodetaric.excel import exporter, reader
from barcodetaric.models import Client, ClientItem


def test_read_codebook_with_headers(tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Barcode", "Περιγραφή", "TARIC"])
    ws.append(["5201234567890", "Καφές αλεσμένος", "0901"])
    ws.append(["5209876543210", "Εμφιαλωμένο νερό", ""])
    path = tmp_path / "codebook.xlsx"
    wb.save(path)

    rows = reader.read_codebook(path)
    assert len(rows) == 2
    assert rows[0].barcode == "5201234567890"
    assert "Καφές" in rows[0].description
    assert rows[0].taric_code == "0901"


def test_read_codebook_no_headers(tmp_path):
    path = tmp_path / "raw.csv"
    path.write_text("5201234567890,Καφές\n5209876543210,Νερό\n", encoding="utf-8")
    rows = reader.read_codebook(path)
    assert len(rows) == 2
    assert all(r.barcode and r.description for r in rows)


def test_export_roundtrip(tmp_path):
    cid = repo.create_client(Client(name="Export Test"))
    repo.upsert_client_item(ClientItem(client_id=cid, barcode="5201234567890",
                                       description_el="Καφές", taric_code="0901", hs4="0901",
                                       taric_source="fts", confidence=0.5))
    out = tmp_path / "out.xlsx"
    n = exporter.export(cid, out)
    assert n == 1 and out.is_file()

    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb.active
    assert ws.max_row == 2  # header + 1
    assert ws.cell(1, 1).value == "Barcode"


def test_preview_columns_autodetect(tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["EAN", "Ονομασία", "Δασμ.", "Εσωτ.Κωδικός"])
    ws.append(["5201219046055", "Nescafe στιγμιαίος 50g", "", "A-100"])
    path = tmp_path / "in.xlsx"
    wb.save(path)

    prev = reader.preview_columns(path)
    assert prev.has_header and prev.n_cols == 4
    assert prev.suggested["barcode"] == 0
    assert prev.suggested["description"] == 1


def test_read_with_mapping_keeps_extra(tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["EAN", "Ονομασία", "Εσωτ.Κωδικός", "Τιμή"])
    ws.append(["5201219046055", "Nescafe στιγμιαίος 50g", "A-100", "9,90"])
    path = tmp_path / "in.xlsx"
    wb.save(path)

    mapping = {"barcode": 0, "description": 1, "taric": None}
    rows = reader.read_with_mapping(path, mapping, extra_cols=[2, 3], has_header=True)
    assert len(rows) == 1
    assert rows[0].extra == {"Εσωτ.Κωδικός": "A-100", "Τιμή": "9,90"}


def test_export_includes_extra_columns(tmp_path):
    import json
    cid = repo.create_client(Client(name="Extra Cols"))
    repo.bulk_upsert_client_items([ClientItem(
        client_id=cid, barcode="5201219046055", description_el="Nescafe",
        source="excel", extra=json.dumps({"Εσωτ.Κωδικός": "A-100"}, ensure_ascii=False))])
    out = tmp_path / "out.xlsx"
    exporter.export(cid, out, include_extra=True)

    from openpyxl import load_workbook
    ws = load_workbook(out).active
    headers = [c.value for c in ws[1]]
    assert "Εσωτ.Κωδικός" in headers
    assert ws.cell(2, headers.index("Εσωτ.Κωδικός") + 1).value == "A-100"

    # χωρίς extra -> μόνο οι βασικές στήλες
    out2 = tmp_path / "out2.xlsx"
    exporter.export(cid, out2, include_extra=False)
    headers2 = [c.value for c in load_workbook(out2).active[1]]
    assert "Εσωτ.Κωδικός" not in headers2


def test_backup_db_creates_file():
    from barcodetaric import db
    dest = db.backup_db("unittest")
    assert dest is not None and dest.is_file()
    assert dest.parent.name == "backups"
